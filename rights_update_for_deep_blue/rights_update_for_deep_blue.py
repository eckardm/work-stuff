from openpyxl import load_workbook, Workbook

revised_copyright_statement_mhc = "Donor(s) have transferred any applicable copyright to the Regents of the University of Michigan but the collection may contain third-party materials for which copyright was not transferred. Patrons are responsible for determining the appropriate use or reuse of materials."
revised_copyright_statement_ua = "Copyright is held by the Regents of the University of Michigan but the collection may contain third-party materials for which copyright is not held. Patrons are responsible for determining the appropriate use or reuse of materials."
revised_negative_copyright_statement_mhc = "Copyright has not been transferred to the Regents of the University of Michigan. Patrons are responsible for determining the appropriate use or reuse of materials."
revised_negative_copyright_statement_ua = "Copyright is not held by the Regents of the University of Michigan. Patrons are responsible for determining the appropriate use or reuse of materials."

wb = load_workbook(filename="RightsForBL.xlsx", read_only=True)
ws = wb.active

rows = []

for row in ws.iter_rows(row_offset=1):
    count = row[0].value
    handle = row[1].value
    collname = row[2].value
    rights = row[3].value
    revised_copyright_statement = ""
    
    mhc = True
    if "(University of Michigan)" in collname:
        mhc = False
    
    print rights
    
    if rights == "Copyright has been transferred to the Regents of the University of Michigan":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
    
    elif rights == "Copyright has been transferred to the Regents of the University of Michigan.":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
      
    elif rights == "Copyright has been transferred to the Regents of the University of Michigan. Researchers are responsible for determining appropriate rights holder.":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
   
    elif rights == "Copyright has been transferred to the Regents of the Unversity of Michigan.":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
    
    elif rights == "Copyright has not been transferred to the Regents of the University of Michigan.":
        if mhc == True:
            revised_copyright_statement = revised_negative_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_negative_copyright_statement_ua
    
    elif rights == "Copyright has not been transferred to the Regents of the University of Michigan. Copyright is held by the Arnold Weinstein Intellectual Property Trust or other third parties. The Works are protected by copyright and permissions to use the Works must be obtained from the copyright owner for any uses of the Works other than for educational or research purposes. Users of the Works may not (i) remove, obscure or modify any copyright notices or other notices included in the Works, or (ii) use, or authorize the use of, Works in any manner that would infringe the copyright thereon.":
        revised_copyright_statement = rights
    
    elif rights == "Copyright has not been transferred to the Regents of the University of Michigan. Researchers are responsible for determining appropriate rights holder.":
        if mhc == True:
            revised_copyright_statement = revised_negative_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_negative_copyright_statement_ua
    
    elif rights == "Copyright held by the Regents of the University of Michigan":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
    
    elif rights == "Copyright held jointly by Constance Cumbey and the Regents of the University of Michigan.":
        revised_copyright_statement = rights
    
    elif rights == "Copyright is held by StoryCorps, Inc., 80 Hanson Place, Brooklyn, NY 11217":
        revised_copyright_statement = rights
    
    elif rights == "Copyright is held by The North Woods Call LLC":
        revised_copyright_statement = "Copyright is held by The North Woods Call, LLC."
    
    elif rights == "Copyright is held by The North Woods Call LLC.":
        revised_copyright_statement = "Copyright is held by The North Woods Call, LLC."
    
    elif rights == "Copyright is held by the Regents of the University of Michigan":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
    
    elif rights == "Copyright is held by the Regents of the University of Michigan.":
        if mhc == True:
            revised_copyright_statement = revised_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_copyright_statement_ua
    
    elif rights == "Copyright is held by the Women Trailblazers in Law Project.":
        revised_copyright_statement = rights
    
    elif rights == "Copyright is not held by the Regents of the University of Michigan.":
        if mhc == True:
            revised_copyright_statement = revised_negative_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_negative_copyright_statement_ua
    
    elif rights == "Copyright not transferred to the Regents of the University of Michigan.":
        if mhc == True:
            revised_copyright_statement = revised_negative_copyright_statement_mhc
        else:
            revised_copyright_statement = revised_negative_copyright_statement_ua
    
    elif rights == "David Littmann transferred his copyright to the Regents of the University of Michigan.":
        revised_copyright_statement = rights
        
    rows.append([count, handle, collname, rights, revised_copyright_statement])
    
wb = Workbook(write_only=True)
ws = wb.create_sheet()

ws.append(["count", "handle", "collname", "rights", "revised_copyright_statement"])
for row in rows:
    ws.append(row)
    
wb.save("RightsForBL-REVISED.xlsx")
