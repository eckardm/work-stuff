import os
from openpyxl import load_workbook
from collections import Counter
from pprint import pprint
from lxml import etree
from itertools import izip_longest
from time import strftime
import shutil

'''
prep of data load into Deep Blue
================================

This is the directory structure that deep blue expect when I ingest items.  

I usually call the dir to ingest archive ( but any name is fine).  Under this directory you want to put every item you want in a separate directory ( one directory per item ).  Give these dirs a unique numeric value.  In each of these dirs.  You need the following files:

dublin_core.xml => This file contains all the metadata. Here is an example.

<dcvalue element="title" qualifier="none">The name of title</dcvalue>

license.txt => this is a text file with the license.  This is a constant.

contents => this file contains a list of all the files to upload for this item ( except for the dublin_core.xml file).  So for example:
license.txt

a.zip<tab>description:This file is really important.  Access restricted to UM users.

b.zip<tab>description:This file is really important.  Access restricted to UM users.

Things to keep in mind:
(1)  All the items in the archive directory should have the same rights - all free, all restricted to UM, all restricted to Bentley IP, etc...

(2) The content file should always list the license.txt file and the other files, if they have a description you need a tab between the filename and the word description:, also you need to include some text that contains any restrictions.  This is to get the icon to show up on the filename.

Access restricted to U-M

Access restricted to Bentley

(3) When creating the dubline_core.xml file, you'll need to know what the dc values are are for deepblue.  Here is a comprehensive list of all of them.  You'll use very few of them.  Note that things like mime type you wont what in the list, since deep blue computes that on ingest.  You always need:

These are for the browse.
title.none
date.issued          
contributor.author

Here is the list:
contributor.none
 contributor.advisor
 contributor.editor
 contributor.author
 contributor.illustrator
 contributor.other
 creator.none
 date.available
 date.copyright
 date.created
 date.issued
 subject.hlbsecondlevel
 identifier.govdoc
 identifier.isbn
 identifier.issn
 identifier.sici
 identifier.ismn
 identifier.other
 identifier.uri
 description.none
 description.abstract
 identifier.orcid
 language.iso
 publisher.none
 date.open
 rights.none
 source.none
 contributor.committeemember
 subject.none
 identifier.citation
 subject.other
 title.none
 type.none
 subject.hlbtoplevel
 identifier.pmid
 identifier.doi
 identifier.source
 identifier.citedreference
 rights.access
 identifier.none
 rights.copyright
 abstract.none
 creator.none
 date.none
 identifier.none
 issued.none
 publisher.none
 rights.none
 subject.none
 title.none
 type.none

(4) If you have items that are actually embargoed.  This means you don't want access to this item at all. You need this in the dublin_core.xml file.  In this example should not be made available for 12 months.

<dcvalue element="date" qualifier="available">WITHHELD_12_MONTHS</dcvalue>

the groups and their rights
===========================

Anonymous - bitstream is free to everyone

Bentley Only Users - this one is strictly restricted to this ip 141.211.39.* without cosign option. 
AND 
must have this in the file description:  Access restricted to Bentley

example:
https://deepblue.lib.umich.edu/handle/2027.42/109255

Bentley Users - This one has more ips at bentley and users can cosign in to get it.

example:
https://deepblue.lib.umich.edu/handle/2027.42/102530

UM Users - access restricted to UM IP address and cosign.

example:
https://deepblue.lib.umich.edu/handle/2027.42/63650'''

# preliminaries
def get_deposit_id():
    while True:
        deposit_id = raw_input("Deposit ID: ")
        if deposit_id in os.listdir("X:\deepblue"):
            break
        print "Enter a valid Deposit ID."
        
    return deposit_id
    
deposit_id = get_deposit_id()

source_directory = os.path.join("X:\deepblue", deposit_id)
temporary_directory = "archive_directory"
target_directory = os.path.join("S:\MLibrary\DeepBlue", deposit_id)

metadata = [filename for filename in os.listdir(source_directory) if filename.startswith("deepBlue_")][0]

# basic metadata check
def get_dc_titles_and_dc_description_abstracts(directory, metadata):
    dc_titles = []
    dc_description_abstracts = []

    wb = load_workbook(filename=os.path.join(directory, metadata), read_only=True, use_iterators=True)
    ws = wb.active

    for row in ws.iter_rows(row_offset=1):
        dc_titles.append(row[1].value)
        dc_description_abstracts.append(row[2].value)
        
    return dc_titles, dc_description_abstracts
    
def get_filenames_and_dc_title_filenames(directory, metadata):
    filenames = []
    dc_title_filenames = []

    for filename in os.listdir(directory):
        if filename != "Thumbs.db" and not filename.startswith("deepBlue_"):
            filenames.append(filename)
            
    wb = load_workbook(filename=os.path.join(directory, metadata), read_only=True, use_iterators=True)
    ws = wb.active

    for row in ws.iter_rows(row_offset=1):
        for title in row[8].value.split("|"):
            dc_title_filenames.append(title)
        
    return filenames, dc_title_filenames

def check_that_dc_titles_are_unique(dc_titles):
    if len(dc_titles) != len(set(dc_titles)):
        print "All titles not unique..."
        print pprint(Counter(dc_titles).most_common())
        quit()

def check_that_dc_description_abstracts_are_unique(dc_description_abstracts):
    if len(dc_description_abstracts) != len(set(dc_description_abstracts)):
        print "All descriptions not unique..."   
        print pprint(Counter(dc_description_abstracts).most_common())
        quit()

def check_that_filenames_match_dc_title_filenames(filenames, dc_title_filenames):
    filenames_not_in_filenames_in_metadata = []

    for filename in filenames:
        if filename not in dc_title_filenames:
            filenames_not_in_filenames_in_metadata.append(filename)
            
    if len(filenames_not_in_filenames_in_metadata) > 0:
        print "All filenames and filenames in metadata do not match..."
        for filename in filenames_not_in_filenames_in_metadata:
            print filename
        quit()
        
def basic_metadata_check(directory, metadata):
    print "Performing basic metadata check..."
    
    dc_titles, dc_description_abstracts = get_dc_titles_and_dc_description_abstracts(directory, metadata)
    check_that_dc_titles_are_unique(dc_titles)
    check_that_dc_description_abstracts_are_unique(dc_description_abstracts)
    
    filenames, dc_title_filenames = get_filenames_and_dc_title_filenames(directory, metadata)
    check_that_filenames_match_dc_title_filenames(filenames, dc_title_filenames)
    
basic_metadata_check(source_directory, metadata)

# make working copy
def make_working_copy(source_directory):
    print "Making working copy..."
    
    os.makedirs(deposit_id)
    
    os.putenv("SOURCE_DIRECTORY", source_directory)    
    os.putenv("TARGET_DIRECTORY", os.path.dirname(os.path.abspath(__file__)))
    os.system('"C:\Program Files\TeraCopy\TeraCopy.exe" Copy %SOURCE_DIRECTORY% %TARGET_DIRECTORY%')
    
make_working_copy(source_directory)
    
# make archive directory
def make_archive_directory(directory):
    os.rename(deposit_id, directory)

def make_item(directory, counter):
    item = "item_"
    number = str(counter).zfill(3)
    item = item + number
    
    os.makedirs(os.path.join(directory, item))
    
    return item    
    
def make_dublin_core(directory, row, item):
    dublin_core = etree.Element("dublin_core")
    
    identifier_other = row[0].value
    if identifier_other:
        etree.SubElement(dublin_core, "dcvalue", element="identifier", qualifier="other").text = identifier_other
    
    dc_title = row[1].value
    etree.SubElement(dublin_core, "dcvalue", element="title", qualifier="none").text = dc_title
    
    dc_description_abstract = row[2].value
    if dc_description_abstract:
        etree.SubElement(dublin_core, "dcvalue", element="description", qualifier="abstract").text = dc_description_abstract
    
    dc_contributor_authors = row[3].value.split("|")
    for dc_contributor_author in dc_contributor_authors:
        etree.SubElement(dublin_core, "dcvalue", element="contributor", qualifier="author").text = dc_contributor_author
        
    if row[4].value:
        dc_contributor_others = row[4].value.split("|")
        for dc_contributor_other in dc_contributor_authors:
            etree.SubElement(dublin_core, "dcvalue", element="contributor", qualifier="other").text = dc_contributor_other
        
    dc_date_issued = str(row[5].value)
    etree.SubElement(dublin_core, "dcvalue", element="date", qualifier="issued").text = dc_date_issued
    
    dc_date_created = str(row[6].value)
    if dc_date_created:
        etree.SubElement(dublin_core, "dcvalue", element="date", qualifier="created").text = dc_date_created
        
    dc_coverage_temporal = row[7].value
    if dc_coverage_temporal:
        etree.SubElement(dublin_core, "dcvalue", element="coverage", qualifier="temporal").text = dc_coverage_temporal
    
    if row[10].value:
        dc_types = row[10].value.split("| ")
        for dc_type in dc_types:
            etree.SubElement(dublin_core, "dcvalue", element="type", qualifier="none").text = dc_type

    dc_rights_access = row[11].value
    if dc_rights_access:
        etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="access").text = dc_rights_access
    
    dc_date_open = row[12].value.strftime("%Y-%m-%d")
    if dc_date_open:
        etree.SubElement(dublin_core, "dcvalue", element="date", qualifier="open").text = dc_date_open
    
    dc_rights_copyright = row[13].value
    if dc_rights_copyright:
        etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="copyright").text = dc_rights_copyright
    
    dc_language_iso = row[14].value
    if dc_language_iso:
        etree.SubElement(dublin_core, "dcvalue", element="language", qualifier="iso").text = dc_language_iso
    
    dublin_core = etree.tostring(dublin_core, pretty_print=True, xml_declaration=True, encoding="utf-8", standalone=False)
    
    with open(os.path.join(directory, item, "dublin_core.xml"), mode="w") as f:
        f.write(dublin_core)
        
def make_license(directory, item):
    with open(os.path.join(directory, item, "license.txt"), mode="w") as f:
        
        f.write("As the designated coordinator for this Deep Blue Collection, \
        I am authorized by the Community members to serve as their representative \
        in all dealings with the Repository. As the designee, I ensure that I \
        have read the Deep Blue policies. Furthermore, I have conveyed to the \
        community the terms and conditions outlined in those policies, including \
        the language of the standard deposit license quoted below and that the \
        community members have granted me the authority to deposit content on \
        their behalf.")
        f.write("\n")
        
def get_dc_title_filenames_and_dc_description_filenames(row):
    dc_title_filenames = row[8].value.split("|")
    
    dc_description_filenames = []
    if row[9].value:
        dc_description_filenames = row[9].value.split("|")
        
    return dc_title_filenames, dc_description_filenames
    
def get_dc_rights_access(row):
    dc_rights_access = row[11].value

    return dc_rights_access
        
def make_contents(directory, item, dc_title_filenames, dc_description_filenames, dc_rights_access):
    with open(os.path.join(directory, item, "contents"), mode="w") as f:
        
        f.write("license.txt\n")
        
        for dc_title_filename, dc_description_filename in izip_longest(dc_title_filenames, dc_description_filenames):
            
            f.write(dc_title_filename)
            
            if dc_description_filename:
                f.write("\tdescription:" + dc_description_filename)
            
            if dc_rights_access.startswith("This content is open for research"):
                f.write("\tpermissions:Anonymous")
            elif dc_rights_access.startswith("Reading room access only"):
                f.write(" Access restricted to Bentley.")
                f.write("\tpermissions:Bentley Only Users")
            else:
                print "Check out the permissions on these... looks like they're complicated."
                print dc_rights_access
                quit()
                
            f.write("\n")
    
def move_objects(directory, item, dc_title_filenames):
    objects = [filename for filename in os.listdir(directory) if filename != metadata and not filename.startswith("item_")]
    
    for object in objects:
        if object in dc_title_filenames:
            
            os.putenv("SOURCE_DIRECTORY", os.path.join(directory, object))
            os.putenv("TARGET_DIRECTORY", os.path.join(directory, item))
            os.system('"C:\Program Files\TeraCopy\TeraCopy.exe" Move %SOURCE_DIRECTORY% %TARGET_DIRECTORY%')

def delete_metadata(directory, metadata):
    os.remove(os.path.join(directory, metadata))

def make_dspace_simple_archive_format(directory, metadata):
    print "Converting AutoPro output to DSpace Simple Archive Format (SAF)..."
    
    make_archive_directory(directory)
    
    wb = load_workbook(filename=os.path.join(directory, metadata), read_only=True, use_iterators=True)
    ws = wb.active

    counter = 1

    for row in ws.iter_rows(row_offset=1):
        item = make_item(directory, counter)
        make_dublin_core(directory, row, item)
        
        make_license(directory, item)
        
        dc_title_filenames, dc_description_filenames = get_dc_title_filenames_and_dc_description_filenames(row)
        dc_rights_access = get_dc_rights_access(row)
        make_contents(directory, item, dc_title_filenames, dc_description_filenames, dc_rights_access)
               
        move_objects(directory, item, dc_title_filenames)
        
        counter += 1
        
    wb._archive.close()
    delete_metadata(directory, metadata)
        
make_dspace_simple_archive_format(temporary_directory, metadata)

# moving temporary directory
def move_to_mlibrary_deep_blue(temporary_directory, target_directory):
    print "Moving to S:\MLibrary\DeepBlue..."
    
    os.makedirs(target_directory)

    os.putenv("SOURCE_DIRECTORY", os.path.join(os.path.dirname(os.path.abspath(__file__)), temporary_directory))
    os.putenv("TARGET_DIRECTORY", target_directory)
    os.system('"C:\Program Files\TeraCopy\TeraCopy.exe" Move %SOURCE_DIRECTORY% %TARGET_DIRECTORY%')
    
move_to_mlibrary_deep_blue(temporary_directory, target_directory)

# cleaning everything up
def clean_up_temporary_directory(directory):
    print "Clean up, clean up, everybody do your share..."
    
    shutil.rmtree(directory)
    
clean_up_temporary_directory(temporary_directory)

print "Alright, we're done!"
