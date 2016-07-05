import os
from openpyxl import load_workbook
from collections import Counter
from lxml import etree
from itertools import izip_longest
from time import strftime
import shutil

# preliminaries
def get_deposit_id():
    while True:
        deposit_id = raw_input("Deposit ID: ")
        if deposit_id in os.listdir("X:\deepblue"):
            break
        print "Enter a valid Deposit ID."
        
    return deposit_id
    
def get_handle():
    handle = raw_input("Handle: ")
    if not handle.startswith("2027.42/"):
        handle = "2027.42/" + handle
        
    return handle
    
deposit_id = get_deposit_id()
handle = get_handle()

source_directory = os.path.join("X:", "deepblue", deposit_id)
temporary_directory = "archive_directory"
target_directory = os.path.join("R:", "MLibrary Drop", "DeepBlue")
beal_directory = os.path.join("X:", "beal", deposit_id)

bentleystaff_items = []

metadata = [filename for filename in os.listdir(source_directory) if filename.startswith("deepBlue_")][0]
logs = [filename for filename in os.listdir(beal_directory) if filename.endswith("zip")][0]

# basic metadata check
def get_dc_titles_and_dc_description_abstracts(directory, metadata):
    dc_titles = []
    dc_description_abstracts = []

    wb = load_workbook(filename=os.path.join(directory, metadata), read_only=True, use_iterators=True)
    ws = wb.active

    for row in ws.iter_rows(row_offset=1):
        dc_titles.append(row[1].value)
        dc_description_abstracts.append(row[2].value)
        
    return dc_titles, dc_description_abstracts
    
def get_filenames_and_dc_title_filenames(directory, metadata):
    filenames = []
    dc_title_filenames = []

    for filename in os.listdir(directory):
        if filename != "Thumbs.db" and not filename.startswith("deepBlue_"):
            filenames.append(filename)
            
    wb = load_workbook(filename=os.path.join(directory, metadata), read_only=True, use_iterators=True)
    ws = wb.active

    for row in ws.iter_rows(row_offset=1):
        for title in row[8].value.split("|"):
            dc_title_filenames.append(title)
        
    return filenames, dc_title_filenames

def check_that_dc_titles_are_unique(dc_titles):
    if len(dc_titles) != len(set(dc_titles)):
        print "These titles are not unique..."
        errors = [dc_title for dc_title, counter in Counter(dc_titles).most_common() if counter > 1]
        for error in errors:
            print error
        quit()

def check_that_dc_description_abstracts_are_unique(dc_description_abstracts):
    if len(dc_description_abstracts) != len(set(dc_description_abstracts)):
        print "These descriptions are not unique..."
        errors = [dc_description_abstract for dc_description_abstract, counter in Counter(dc_description_abstracts).most_common() if counter > 1]
        for error in errors:
            print error
        quit()

def check_that_filenames_match_dc_title_filenames(filenames, dc_title_filenames):
    filenames_not_in_filenames_in_metadata = []

    for filename in filenames:
        if filename not in dc_title_filenames:
            filenames_not_in_filenames_in_metadata.append(filename)
            
    if len(filenames_not_in_filenames_in_metadata) > 0:
        print "These filenames are not in the metadata..."
        for filename in filenames_not_in_filenames_in_metadata:
            print filename
        quit()
        
def basic_metadata_check(directory, metadata):
    print "Performing basic metadata check..."
    
    dc_titles, dc_description_abstracts = get_dc_titles_and_dc_description_abstracts(directory, metadata)
    check_that_dc_titles_are_unique(dc_titles)
    check_that_dc_description_abstracts_are_unique(dc_description_abstracts)
    
    filenames, dc_title_filenames = get_filenames_and_dc_title_filenames(directory, metadata)
    check_that_filenames_match_dc_title_filenames(filenames, dc_title_filenames)
    
basic_metadata_check(source_directory, metadata)

# make working copy
def make_working_copy(source_directory, beal_directory, logs):
    print "Making working copy..."
    
    shutil.copytree(source_directory, os.path.join(os.path.dirname(os.path.abspath(__file__)), deposit_id))
    shutil.copy(os.path.join(beal_directory, logs), os.path.join(os.path.dirname(os.path.abspath(__file__)), deposit_id))
    
make_working_copy(source_directory, beal_directory, logs)
    
# make archive directory
def make_archive_directory(directory):
    os.rename(deposit_id, directory)

def make_item(directory, counter):
    item = "item_"
    number = str(counter).zfill(3)
    item = item + number
    
    os.makedirs(os.path.join(directory, item))
    
    return item    
    
def make_dublin_core(directory, row, item):
    dublin_core = etree.Element("dublin_core")
    
    identifier_other = row[0].value
    if identifier_other:
        etree.SubElement(dublin_core, "dcvalue", element="identifier", qualifier="other").text = identifier_other
    
    dc_title = row[1].value.split(" - ")[-1]
    etree.SubElement(dublin_core, "dcvalue", element="title", qualifier="none").text = dc_title
    
    dc_relation_ispartof = " - ".join(row[1].value.split(" - ")[:-1])
    if dc_relation_ispartof:
        etree.SubElement(dublin_core, "dcvalue", element="relation", qualifier="ispartof").text = dc_relation_ispartof   
    
    dc_description_abstract = row[2].value
    if dc_description_abstract:
        etree.SubElement(dublin_core, "dcvalue", element="description", qualifier="abstract").text = dc_description_abstract
    
    dc_contributor_authors = row[3].value.split("|")
    for dc_contributor_author in dc_contributor_authors:
        if dc_contributor_author.startswith("University of Michigan."):
            dc_contributor_author = dc_contributor_author.replace("University of Michigan.", "University of Michigan,")
        etree.SubElement(dublin_core, "dcvalue", element="contributor", qualifier="author").text = dc_contributor_author
        
    if row[4].value:
        dc_contributor_others = row[4].value.split("|")
        for dc_contributor_other in dc_contributor_authors:
            etree.SubElement(dublin_core, "dcvalue", element="contributor", qualifier="other").text = dc_contributor_other
        
    dc_date_issued = str(row[5].value)
    etree.SubElement(dublin_core, "dcvalue", element="date", qualifier="issued").text = dc_date_issued
    
    dc_date_created = str(row[6].value)
    if dc_date_created:
        etree.SubElement(dublin_core, "dcvalue", element="date", qualifier="created").text = dc_date_created
        
    dc_coverage_temporal = str(row[7].value)
    if dc_coverage_temporal:
        etree.SubElement(dublin_core, "dcvalue", element="coverage", qualifier="temporal").text = dc_coverage_temporal
    
    if row[10].value:
        dc_types = row[10].value.split("| ")
        for dc_type in dc_types:
            etree.SubElement(dublin_core, "dcvalue", element="type", qualifier="none").text = dc_type

    dc_rights_access = row[11].value
    if dc_rights_access:
        etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="access").text = dc_rights_access
        if dc_rights_access.startswith("Executive Records") or dc_rights_access.startswith("Personnel Records") or dc_rights_access.startswith("Student Records") or dc_rights_access.startswith("Patient/Client Records"):
            etree.SubElement(dublin_core, "dcvalue", element="description", qualifier="restriction").text = "RESTRICTED"
            bentleystaff_items.append(item)
    
    dc_date_open = row[12].value
    if dc_date_open:
        etree.SubElement(dublin_core, "dcvalue", element="date", qualifier="open").text = dc_date_open.strftime("%Y-%m-%d")
    
    dc_rights_copyright = row[13].value
    if dc_rights_copyright:
        if dc_rights_copyright == "Copyright has been transferred to the Regents of the University of Michigan.":
            if dc_contributor_author.startswith("University of Michigan"):
                etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="copyright").text = "Copyright is held by the Regents of the University of Michigan but the collection may contain third-party materials for which copyright is not held. Patrons are responsible for determining the appropriate use or reuse of materials."
            else:
                etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="copyright").text = "Donor(s) have transferred any applicable copyright to the Regents of the University of Michigan but the collection may contain third-party materials for which copyright was not transferred. Patrons are responsible for determining the appropriate use or reuse of materials."
        else:
            if dc_contributor_author.startswith("University of Michigan"):
                etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="copyright").text = "Copyright is not held by the Regents of the University of Michigan. Patrons are responsible for determining the appropriate use or reuse of materials."
            else:
                etree.SubElement(dublin_core, "dcvalue", element="rights", qualifier="copyright").text= "Copyright has not been transferred to the Regents of the University of Michigan. Patrons are responsible for determining the appropriate use or reuse of materials."
    
    dc_language_iso = row[14].value
    if dc_language_iso:
        etree.SubElement(dublin_core, "dcvalue", element="language", qualifier="iso").text = dc_language_iso
    
    dublin_core = etree.tostring(dublin_core, pretty_print=True, xml_declaration=True, encoding="utf-8", standalone=False)
    
    with open(os.path.join(directory, item, "dublin_core.xml"), mode="w") as f:
        f.write(dublin_core)
        
def make_license(directory, item):
    with open(os.path.join(directory, item, "license.txt"), mode="w") as f:
        
        f.write("As the designated coordinator for this Deep Blue Collection, \
        I am authorized by the Community members to serve as their representative \
        in all dealings with the Repository. As the designee, I ensure that I \
        have read the Deep Blue policies. Furthermore, I have conveyed to the \
        community the terms and conditions outlined in those policies, including \
        the language of the standard deposit license quoted below and that the \
        community members have granted me the authority to deposit content on \
        their behalf.")
        f.write("\n")
        
def get_dc_title_filenames_and_dc_description_filenames(row):
    dc_title_filenames = row[8].value.split("|")
    
    dc_description_filenames = []
    if row[9].value:
        dc_description_filenames = row[9].value.split("|")
        
    return dc_title_filenames, dc_description_filenames
    
def get_dc_rights_access(row):
    dc_rights_access = row[11].value

    return dc_rights_access
        
def make_contents(directory, item, dc_title_filenames, dc_description_filenames, dc_rights_access):
    with open(os.path.join(directory, item, "contents"), mode="w") as f:
        
        f.write("license.txt\n")
        f.write(logs + "\tdescription:Administrative information. Access restricted to Bentley staff. \tpermissions:-r 'BentleyStaff'\n")
        
        for dc_title_filename, dc_description_filename in izip_longest(dc_title_filenames, dc_description_filenames):
            
            f.write(dc_title_filename)
            
            if dc_description_filename:
                f.write("\tdescription:" + dc_description_filename)
            
            if dc_rights_access.startswith("This content is open for research"):
                f.write("\tpermissions:-r 'Anonymous'")
            
            elif dc_rights_access.startswith("Reading room access only"):
                if not dc_description_filename:
                    f.write("\tdescription:Access restricted to Bentley Reading Room.")
                else:
                    if dc_description_filename.endswith("."):
                        f.write(" Access restricted to Bentley Reading Room.")
                    else: 
                        f.write(". Access restricted to Bentley Reading Room.")
                f.write("\tpermissions:-r 'Bentley Only Users'")
                
            elif dc_rights_access.startswith("Executive Records") or dc_rights_access.startswith("Personnel Records") or dc_rights_access.startswith("Student Records") or dc_rights_access.startswith("Patient/Client Records"):
                if not dc_description_filename:
                    f.write("\tdescription:Access restricted to Bentley staff.")
                else:
                    if dc_description_filename.endswith("."):
                        f.write(" Access restricted to Bentley staff.")
                    else: 
                        f.write(". Access restricted to Bentley staff.")
                f.write("\tpermissions:-r 'BentleyStaff'")
                
            else:
                print "Wierd permissions on " + item + ", DEAL WITH IT!"

            f.write("\n")
    
def copy_logs(directory, logs, item):
    shutil.copy(os.path.join(directory, logs), os.path.join(directory, item))
    
def move_objects(directory, item, dc_title_filenames):
    objects = [filename for filename in os.listdir(directory) if filename != metadata and not filename.startswith("item_")]
    
    for object in objects:
        if object in dc_title_filenames:
            
            shutil.move(os.path.join(directory, object), os.path.join(directory, item))

def delete_metadata(directory, metadata):
    os.remove(os.path.join(directory, metadata))

def make_dspace_simple_archive_format(directory, metadata):
    print "Converting AutoPro output to DSpace Simple Archive Format (SAF)..."
    
    make_archive_directory(directory)
    
    wb = load_workbook(filename=os.path.join(directory, metadata), read_only=True, use_iterators=True)
    ws = wb.active

    counter = 1

    for row in ws.iter_rows(row_offset=1):
        item = make_item(directory, counter)
        make_dublin_core(directory, row, item)
        
        make_license(directory, item)
        
        copy_logs(directory, logs, item)
        
        dc_title_filenames, dc_description_filenames = get_dc_title_filenames_and_dc_description_filenames(row)
        dc_rights_access = get_dc_rights_access(row)
        make_contents(directory, item, dc_title_filenames, dc_description_filenames, dc_rights_access)
               
        move_objects(directory, item, dc_title_filenames)
        
        counter += 1
        
    wb._archive.close()
    delete_metadata(directory, metadata)
        
make_dspace_simple_archive_format(temporary_directory, metadata)

# moving temporary directory
def move_to_mlibrary_deep_blue(temporary_directory, target_directory, deposit_id):
    print "Moving to " + target_directory + "..."
    
    shutil.copytree(os.path.join(os.path.dirname(os.path.abspath(__file__)), temporary_directory), os.path.join(target_directory, deposit_id))
    os.remove(os.path.join(target_directory, deposit_id, logs))
    
move_to_mlibrary_deep_blue(temporary_directory, target_directory, deposit_id)

# restructuring bentleystaff items
def restructuring_bentleystaff_items(target_directory, bentleystaff_items):
    if len(bentleystaff_items) > 0:
        print "Restructuring BentleyStaff items..."
    
    for bentleystaff_item in bentleystaff_items:
        shutil.copytree(os.path.join(target_directory, deposit_id, bentleystaff_item), os.path.join(target_directory, deposit_id + "-BentleyStaff", bentleystaff_item))
        shutil.rmtree(os.path.join(target_directory, deposit_id, bentleystaff_item))
        
    if os.path.exists(os.path.join(target_directory, deposit_id + "-BentleyStaff")) == True and len(os.listdir(os.path.join(target_directory, deposit_id + "-BentleyStaff"))) == 0:
        shutil.rmtree(os.path.join(target_directory, deposit_id + "-BentleyStaff"))
    
restructuring_bentleystaff_items(target_directory, bentleystaff_items)    

# rename directories so they're more helpful for jose
def rename_directories_for_jose(target_directory):
    print "Renaming directories so they're more helpful for Jose..."
    
    for dir in os.listdir(target_directory):
        if dir.startswith(deposit_id):
            
            new_dir = new_dir = deposit_id + "_to_" + handle.replace(".", "-").replace("/", "-")
            if dir.endswith("-BentleyStaff"):
                new_dir = new_dir + "-BentleyStaff"
        
            os.rename(os.path.join(target_directory, dir), os.path.join(target_directory, new_dir))
                
rename_directories_for_jose(target_directory)

# cleaning everything up
def clean_up_temporary_directory(directory):
    print "Clean up, clean up, everybody do your share..."
    
    shutil.rmtree(directory)
    
clean_up_temporary_directory(temporary_directory)

print "Alright, we're done!"
